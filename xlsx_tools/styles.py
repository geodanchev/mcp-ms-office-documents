"""Cell styling primitives for the Excel tool.

Deliberately general-purpose: this module gives the caller the means to colour
a cell or apply a named style from their own template, and takes no view on
what those styles should mean. Presentation conventions belong to whoever is
writing the markdown, not to the server.

Two mechanisms, usable together:

**Inline attributes** — ``<!-- styles: B2=bg:yellow;bold -->`` sets a fill,
font colour or weight directly. Good for one-off emphasis.

**Named styles from a template** — ``<!-- styles: B2=style:Total -->`` applies
a style defined in ``custom_xlsx_template.xlsx``. Good for a house look
applied consistently, since the definition lives in the template rather than
in every document's markdown.

Named styles are applied first and inline attributes layer on top, so
``style:Total;bg:yellow`` means "the Total style, but highlighted".
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field

from copy import copy

from openpyxl.styles import Alignment, Border, Font, PatternFill

logger = logging.getLogger(__name__)

# Colour words accepted in place of a hex value. Kept small and unsurprising —
# this is a convenience for hand-written markdown, not a design system. Values
# are mid-tone enough to read against black text.
NAMED_COLORS: dict[str, str] = {
    "black": "000000",
    "white": "FFFFFF",
    "red": "FF0000",
    "darkred": "C00000",
    "orange": "FFA500",
    "yellow": "FFFF00",
    "green": "00B050",
    "darkgreen": "006100",
    "blue": "0070C0",
    "darkblue": "002060",
    "purple": "7030A0",
    "pink": "FFC7CE",
    "brown": "8B4513",
    "grey": "808080",
    "lightgrey": "D9D9D9",
    "darkgrey": "404040",
    "cyan": "00B0F0",
    "magenta": "FF00FF",
}
# Accept the US spelling for every "grey" entry.
NAMED_COLORS.update({
    name.replace("grey", "gray"): value
    for name, value in NAMED_COLORS.items() if "grey" in name
})

_HEX_COLOR_RE = re.compile(r'^#?([0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})$')

# Flag attributes that take no value.
_FLAG_ATTRS = frozenset({"bold", "italic", "underline"})

# Upper bound on cells a single range entry may cover. A styles directive is
# meant for emphasis; anything larger is a mistake (a stray `A1:Z100000`)
# rather than an intent, and expanding it would stall generation.
MAX_STYLED_CELLS = 10_000

# The template filename resolved through template_utils, following the
# `custom_<kind>_template.<ext>` convention the docx/pptx/email tools use.
TEMPLATE_FILENAME = "custom_xlsx_template.xlsx"


@dataclass
class StyleSpec:
    """A parsed style instruction for one cell."""

    named_style: str | None = None
    background: str | None = None   # 6- or 8-digit hex, no leading '#'
    font_color: str | None = None
    bold: bool = False
    italic: bool = False
    underline: bool = False

    def is_empty(self) -> bool:
        return not any((self.named_style, self.background, self.font_color,
                        self.bold, self.italic, self.underline))


def parse_color(value: str) -> str | None:
    """Normalise a colour to the bare hex form openpyxl expects.

    Accepts ``FFFF00``, ``#FFFF00``, the 8-digit ARGB form, and the names in
    :data:`NAMED_COLORS`. Returns None if the value isn't a colour, so the
    caller can warn rather than write a broken fill.
    """
    if not value:
        return None
    text = value.strip()
    match = _HEX_COLOR_RE.match(text)
    if match:
        return match.group(1).upper()
    return NAMED_COLORS.get(text.lower())


def parse_style_spec(spec: str) -> StyleSpec:
    """Parse ``"bg:yellow;bold"`` into a :class:`StyleSpec`.

    Attributes are separated by ``;`` because ``,`` already separates entries
    in the directive and ``:`` already separates an attribute from its value.
    Unknown attributes are logged and skipped rather than failing the
    document — a typo in one colour should not cost the caller the workbook.
    """
    result = StyleSpec()
    for attr in spec.split(';'):
        attr = attr.strip()
        if not attr:
            continue

        key, sep, raw_value = attr.partition(':')
        key = key.strip().lower()
        raw_value = raw_value.strip()

        if not sep:
            if key in _FLAG_ATTRS:
                setattr(result, key, True)
            else:
                logger.warning(
                    "Unknown style attribute '%s' — expected one of %s, or "
                    "bg:/color:/style: with a value.",
                    attr, ", ".join(sorted(_FLAG_ATTRS)),
                )
            continue

        if key == "style":
            result.named_style = raw_value or None
        elif key in ("bg", "background", "fill"):
            color = parse_color(raw_value)
            if color:
                result.background = color
            else:
                logger.warning(
                    "Unrecognised background colour '%s' — use a 6-digit hex "
                    "value (FFFF00) or a name (%s).",
                    raw_value, ", ".join(sorted(NAMED_COLORS)[:6]) + ", …",
                )
        elif key in ("color", "fg", "font"):
            color = parse_color(raw_value)
            if color:
                result.font_color = color
            else:
                logger.warning("Unrecognised font colour '%s'.", raw_value)
        else:
            logger.warning("Unknown style attribute '%s'.", key)

    return result


# A cell target: a column letter plus either an absolute worksheet row (B2) or
# a table-relative row in brackets (B[0]). A styles directive belongs to
# exactly one table, so the bracket form has an unambiguous anchor — unlike in
# a formula, where it means "the row I am on".
_TARGET_RE = re.compile(r'^([A-Za-z]{1,3})(?:(\d+)|\[([+-]?\d+)\])$')


def _resolve_target(token: str, table_start_row: int | None) -> tuple[str, int] | None:
    """Resolve one target token to ``(column_letter, absolute_row)``.

    ``B2`` is an absolute worksheet row. ``B[0]`` is the table's first data
    row, ``B[1]`` the second, and ``B[-1]`` the header — resolved against
    ``table_start_row``, which is the header row of the table this directive
    is attached to.
    """
    match = _TARGET_RE.match(token.strip())
    if not match:
        return None

    column, absolute_row, relative_row = match.groups()
    column = column.upper()

    if absolute_row is not None:
        row = int(absolute_row)
    else:
        if table_start_row is None:
            return None
        # start_row is the header; data row 0 sits directly beneath it.
        row = table_start_row + 1 + int(relative_row)

    return (column, row) if row >= 1 else None


def _expand_target(target: str, table_start_row: int | None) -> list[str]:
    """Expand a target ("B2", "B[0]", "B2:D5") into absolute coordinates.

    Ranges expand to the full rectangle, not just the corners — a caller who
    writes ``B2:D5`` means the block. Oversized ranges are dropped with a
    warning rather than silently truncated to something that looks like it
    worked.
    """
    start_token, sep, end_token = target.partition(':')

    start = _resolve_target(start_token, table_start_row)
    if start is None:
        return []
    if not sep:
        return [f"{start[0]}{start[1]}"]

    end = _resolve_target(end_token, table_start_row)
    if end is None:
        return []

    from openpyxl.utils import column_index_from_string, get_column_letter

    col_start, col_end = sorted(
        (column_index_from_string(start[0]), column_index_from_string(end[0]))
    )
    row_start, row_end = sorted((start[1], end[1]))

    total = (col_end - col_start + 1) * (row_end - row_start + 1)
    if total > MAX_STYLED_CELLS:
        logger.warning(
            "Style range '%s' covers %d cells, over the %d limit — skipped. "
            "Style a smaller block.",
            target, total, MAX_STYLED_CELLS,
        )
        return []

    return [
        f"{get_column_letter(col)}{row}"
        for col in range(col_start, col_end + 1)
        for row in range(row_start, row_end + 1)
    ]


def parse_styles_directive(
    value: str,
    table_start_row: int | None = None,
) -> dict[str, StyleSpec]:
    """Parse a ``styles:`` directive into ``{coordinate: StyleSpec}``.

    Entries are ``<target>=<spec>`` separated by commas::

        B2=bg:yellow, C3:C8=color:red;bold, D2=style:Total

    Later entries win on overlap, so a range can set a baseline and a single
    cell can override one member of it.

    Args:
        value: The directive's text, without the ``styles:`` prefix.
        table_start_row: Header row of the table this directive is attached
            to, used to resolve the ``B[0]`` form. When None, only absolute
            coordinates resolve.

    Returns:
        Absolute worksheet coordinates mapped to their style. Malformed
        entries are logged and skipped.
    """
    if not value:
        return {}

    styles: dict[str, StyleSpec] = {}
    for entry in value.split(','):
        entry = entry.strip()
        if not entry:
            continue

        target, sep, spec_text = entry.partition('=')
        if not sep:
            logger.warning(
                "Malformed styles entry '%s' — expected <cell>=<style>, "
                "e.g. B2=bg:yellow.", entry,
            )
            continue

        coordinates = _expand_target(target.strip(), table_start_row)
        if not coordinates:
            logger.warning(
                "Could not resolve style target '%s' — expected a cell (B2), "
                "a table-relative cell (B[0]) or a range (B2:D5).",
                target.strip(),
            )
            continue

        spec = parse_style_spec(spec_text)
        if spec.is_empty():
            logger.warning("Style entry '%s' set nothing usable.", entry)
            continue

        for coordinate in coordinates:
            styles[coordinate] = spec

    return styles


def _apply_named_style(cell, style_name: str) -> None:
    """Apply a named style without discarding what it has no opinion about.

    An openpyxl ``NamedStyle`` bundles font, fill, border, alignment *and*
    number format, and ``cell.style = name`` replaces all five at once. Left
    alone that silently undoes the table renderer's work: a totals row given
    ``style:Total`` in a ``currency:$`` column loses ``$#,##0.00``, its
    borders and its alignment, because a style defining only a font resets
    those three to their defaults.

    Blanket-restoring all three would be wrong in the other direction — a
    template style that deliberately declares ``0.00%`` must still apply it.
    So each property is restored only where the style left it at its default,
    which is precisely the case where the style expresses no preference. The
    defaults are read off the cell after application, since at that point the
    cell's value *is* the style's value.

    The consequence, which is not solvable rather than merely unsolved: a
    style that deliberately sets ``number_format='General'`` — a "Plain"
    style meant to strip an inherited currency format back to a raw number —
    is treated as having no opinion, so the inherited format wins instead.
    Reading the ``NamedStyle`` object doesn't help; "unset" and "explicitly
    General" are the same value there, and openpyxl serialises both to the
    identical ``<xf numFmtId="0" …/>`` with no ``applyNumberFormat`` flag.
    The information isn't in the file, so no amount of introspection
    recovers it. Clearing an inherited format needs an explicit attribute
    (a future ``format:General``), not a smarter reader.
    """
    previous_number_format = cell.number_format
    previous_border = copy(cell.border)
    previous_alignment = copy(cell.alignment)

    cell.style = style_name

    if cell.number_format == 'General':
        cell.number_format = previous_number_format
    if cell.border == Border():
        cell.border = previous_border
    if cell.alignment == Alignment():
        cell.alignment = previous_alignment


def apply_style_spec(cell, spec: StyleSpec, available_styles: set[str] | None = None) -> None:
    """Apply a :class:`StyleSpec` to a cell.

    A named style is applied first, then inline attributes layer on top and
    win — ``style:Total;bg:yellow`` reads as "the Total style, but
    highlighted". Attributes the spec doesn't mention are left alone, so the
    number format, border and alignment the table renderer set survive (see
    :func:`_apply_named_style` for the one case where a style overrides them
    on purpose).
    """
    if spec.named_style:
        if available_styles is not None and spec.named_style not in available_styles:
            logger.warning(
                "Named style '%s' is not defined in the workbook template "
                "(available: %s) — ignoring.",
                spec.named_style,
                ", ".join(sorted(available_styles)) or "none",
            )
        else:
            try:
                _apply_named_style(cell, spec.named_style)
            except Exception as e:
                logger.warning("Could not apply named style '%s': %s", spec.named_style, e)

    if spec.background:
        cell.fill = PatternFill(
            fill_type="solid", start_color=spec.background, end_color=spec.background
        )

    if spec.font_color or spec.bold or spec.italic or spec.underline:
        current = cell.font
        cell.font = Font(
            name=current.name,
            size=current.size,
            # A flag that isn't set leaves the existing value alone, so
            # `bg:yellow` on an already-bold header doesn't un-bold it.
            bold=True if spec.bold else current.bold,
            italic=True if spec.italic else current.italic,
            underline="single" if spec.underline else current.underline,
            color=spec.font_color or current.color,
        )


# ── Template-defined named styles ────────────────────────────────────────────


@dataclass
class TemplateStyles:
    """Named styles loaded from the optional workbook template."""

    names: set[str] = field(default_factory=set)
    _styles: list = field(default_factory=list)

    def register_into(self, workbook) -> set[str]:
        """Add the template's named styles to ``workbook``; return their names."""
        registered: set[str] = set()
        for style in self._styles:
            if style.name in workbook.named_styles:
                registered.add(style.name)
                continue
            try:
                workbook.add_named_style(style)
                registered.add(style.name)
            except Exception as e:
                logger.warning("Could not register named style '%s': %s", style.name, e)
        return registered


def load_template_styles() -> TemplateStyles:
    """Load named styles from ``custom_xlsx_template.xlsx``, if one exists.

    Resolved through :mod:`template_utils`, so it follows the same
    custom-then-default search path as the docx, pptx and email templates.
    Absence of a template is the normal case and is not an error — it simply
    means no named styles are available to reference.
    """
    try:
        from template_utils import find_file_in_template_dirs
    except ImportError:  # pragma: no cover — defensive
        return TemplateStyles()

    path = find_file_in_template_dirs(TEMPLATE_FILENAME)
    if path is None:
        logger.debug("No %s found; named styles unavailable.", TEMPLATE_FILENAME)
        return TemplateStyles()

    try:
        from openpyxl import load_workbook
        template = load_workbook(str(path))
    except Exception as e:
        logger.warning("Could not read Excel template %s: %s", path, e)
        return TemplateStyles()

    try:
        # `Normal` is openpyxl's built-in default and exists in every
        # workbook; re-registering it is a no-op at best and a conflict at
        # worst, so it is not offered as a referenceable name.
        styles = [s for s in template._named_styles if s.name != "Normal"]
    except Exception as e:
        logger.warning("Could not read named styles from %s: %s", path, e)
        return TemplateStyles()
    finally:
        template.close()

    if styles:
        logger.info(
            "Loaded %d named style(s) from %s: %s",
            len(styles), path.name, ", ".join(s.name for s in styles),
        )
    return TemplateStyles(names={s.name for s in styles}, _styles=styles)
