"""Circular-reference detection for generated XLSX workbooks.

A formula that depends on itself — directly (``A1: =A1+1``) or through a chain
(``A1 → B1 → A1``) — is accepted by openpyxl without complaint. Excel then
shows a warning dialog on open and resolves the cell to 0, so the workbook we
hand the user is quietly wrong. Nothing else in the pipeline catches this.

This module builds a dependency graph from the formula *strings* and runs an
iterative DFS cycle detection over it. It is deliberately self-contained: pure
string and graph work over what openpyxl already parsed, with no formula
evaluation engine and no third-party dependency beyond openpyxl itself.

The detection is best-effort and never raises — a workbook that can't be
analysed is reported as "no cycles" rather than blocking document generation.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from typing import Any

logger = logging.getLogger(__name__)

# Synthetic error type for circular references. Deliberately distinct from the
# seven OOXML error sentinels (#REF!, #DIV/0!, …) so a reader can tell this was
# found by our graph analysis rather than reported by Excel.
CIRCULAR_ERROR_TYPE = "#CIRC!"

# Excel's grid limits, used to reject phantom coordinate matches.
_MAX_EXCEL_ROW = 1_048_576
_MAX_EXCEL_COL = 16_384  # XFD


@dataclass
class CellError:
    """A single problematic cell found by static analysis."""

    sheet: str          # Sheet name as stored in the workbook (original casing).
    coordinate: str     # Excel coordinate, e.g. "B5".
    error_type: str     # CIRCULAR_ERROR_TYPE.

    @property
    def location(self) -> str:
        """``Sheet!Cell`` reference suitable for showing to the caller."""
        return _format_location(self.sheet, self.coordinate)

    def __str__(self) -> str:
        return f"{self.location}: {self.error_type}"


# ── Location keys ────────────────────────────────────────────────────────────


def _format_location(sheet: str, coordinate: str) -> str:
    """Build the ``"Sheet!Cell"`` key used as a graph node identifier."""
    if re.search(r"[^A-Za-z0-9_]", sheet):
        return f"'{sheet}'!{coordinate}"
    return f"{sheet}!{coordinate}"


def _parse_location(location_key: str) -> tuple[str, str]:
    """Parse ``"Sheet!Cell"`` / ``"'Sheet Name'!Cell"`` into (sheet, cell)."""
    if location_key.startswith("'"):
        close = location_key.find("'", 1)
        if close != -1 and location_key[close + 1:close + 2] == "!":
            return location_key[1:close], location_key[close + 2:]
    sheet, _, coordinate = location_key.partition("!")
    return sheet, coordinate


def _build_sheet_lookup(sheet_names: list[str]) -> dict[str, str]:
    """Map uppercased sheet name → the workbook's original casing.

    Formula text may reference a sheet in any casing; Excel matches
    case-insensitively, so we normalise before resolving.
    """
    return {name.upper(): name for name in sheet_names}


# ── Reference extraction ─────────────────────────────────────────────────────

# Pull cell references out of a formula string. Handles cross-sheet
# (``Sheet1!A1``, ``'Sheet 1'!A1``), local (``A1``, ``$A$1``), ranges
# (``A1:B5``) and 3D refs (``Sheet1:Sheet3!A1``). External-workbook and
# structured references are removed before this runs, so there is no
# ``[Workbook.xlsx]`` alternative here.
#
# Regex rather than a real formula parser is the right trade: we need the
# reference set, not an AST. A false positive only adds a spurious graph edge —
# at worst a spurious cycle warning in the logs — never a missed cycle.
_REF_TOKENS_RE = re.compile(
    r"""
    (?:                                 # optional sheet prefix
        '?                              #   optional opening quote
        ([A-Za-z_][\w\s]*)              #   group 1: sheet name
        '?                              #   optional closing quote
        (?:                             #   optional :sheet (3D reference)
            : '? ([A-Za-z_][\w\s]*) '?  #     group 2: end sheet name
        )?
        !                               #   bang
    )?
    (\$?[A-Z]{1,3}\$?\d{1,7})           # group 3: first cell (B2, $A$1)
    (?: : (\$?[A-Z]{1,3}\$?\d{1,7}) )?  # group 4: optional range end
    """,
    re.IGNORECASE | re.VERBOSE,
)


def _normalize_coord(coord: str) -> str:
    """Strip absolute-reference markers: ``$A$1`` → ``A1``."""
    return coord.replace("$", "").upper()


def _col_to_num(letters: str) -> int:
    """Convert Excel column letters to a 1-based column number ('A' → 1)."""
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _num_to_col(n: int) -> str:
    """Convert a 1-based column number to Excel column letters (1 → 'A')."""
    letters = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _is_valid_coord(coord: str) -> bool:
    """Return True if ``coord`` is a real Excel cell coordinate.

    Rejects tokens outside Excel's grid (past XFD1048576). Note this cannot
    catch a coordinate-shaped substring of a longer name — ``BLE1``, from
    ``Table1[…]``, is column 1669 row 1 and perfectly valid in isolation.
    Those are removed by :func:`_strip_unresolvable_refs` before matching.
    """
    match = re.fullmatch(r"([A-Z]{1,3})(\d+)", coord.upper())
    if not match:
        return False
    row = int(match.group(2))
    if not 1 <= row <= _MAX_EXCEL_ROW:
        return False
    return 1 <= _col_to_num(match.group(1)) <= _MAX_EXCEL_COL


def _expand_range(start: str, end: str) -> list[str]:
    """Expand ``('A1','B3')`` into ``['A1','A2','A3','B1','B2','B3']``."""
    start_m = re.match(r"([A-Z]+)(\d+)", start.upper())
    end_m = re.match(r"([A-Z]+)(\d+)", end.upper())
    if not start_m or not end_m:
        return [start]

    c1, r1 = _col_to_num(start_m.group(1)), int(start_m.group(2))
    c2, r2 = _col_to_num(end_m.group(1)), int(end_m.group(2))
    if c1 > c2:
        c1, c2 = c2, c1
    if r1 > r2:
        r1, r2 = r2, r1

    # Cap expansion so a whole-column range (A1:A1048576) can't blow up memory.
    # A cycle that runs through a range this large but touches neither corner
    # is not a case worth paying for.
    if (c2 - c1 + 1) * (r2 - r1 + 1) > 1000:
        return [_normalize_coord(start), _normalize_coord(end)]

    return [
        f"{_num_to_col(c)}{r}"
        for c in range(c1, c2 + 1)
        for r in range(r1, r2 + 1)
    ]


def _strip_string_literals(formula: str) -> str:
    """Remove double-quoted string literals from a formula string.

    Coordinate-shaped text inside a string literal is text, not a reference.
    Without stripping, ``="see cell A1 for context"`` produces a phantom ``A1``
    edge — and a label like ``=A1&" totals"`` placed *in* A1 would be reported
    as a circular reference that doesn't exist.

    Excel escapes a quote inside a literal by doubling it (``""``). Each
    literal is replaced by a single space so tokens either side can't fuse into
    a new coordinate-shaped token.
    """
    out: list[str] = []
    i, n = 0, len(formula)
    while i < n:
        if formula[i] != '"':
            out.append(formula[i])
            i += 1
            continue
        # Inside a literal — skip to the closing unescaped quote.
        i += 1
        while i < n:
            if formula[i] == '"':
                if i + 1 < n and formula[i + 1] == '"':
                    i += 2      # escaped quote, still inside the literal
                    continue
                i += 1          # closing quote
                break
            i += 1
        out.append(" ")
    return "".join(out)


# References this analysis cannot resolve, and which must therefore contribute
# no graph edges at all:
#
#   [Book.xlsx]Sheet1!A1   external workbook — the cell lives in another file
#   Table1[Revenue]        structured reference — no table-name → range mapping
#
# Dropping just the brackets is not enough: the *name* in front of them is
# itself coordinate-shaped. "Table1" contains "ble1" (column 1669, row 1),
# which _is_valid_coord accepts, so the surrounding token has to go too.
#
# Known gap: a quoted external path with non-word characters in the sheet name
# (='[C:\\Reports\\FY24.xlsx]P&L'!B2) isn't matched here. This server never
# generates that form, and the worst case is a spurious edge — never a missed
# cycle — so it isn't worth a more fragile pattern.
_EXTERNAL_REF_RE = re.compile(
    r"\[[^\]]*\]'?[\w\s]*'?!"
    r"\$?[A-Za-z]{1,3}\$?\d{1,7}(?::\$?[A-Za-z]{1,3}\$?\d{1,7})?"
)
_STRUCTURED_REF_RE = re.compile(r"\w+\[[^\]]*\]")


def _strip_unresolvable_refs(formula: str) -> str:
    """Blank out external-workbook and structured references.

    Each is replaced by a space so neighbouring tokens can't fuse into a new
    coordinate-shaped token. External references are removed first — they
    start with ``[`` and so aren't matched by the structured pattern.
    """
    return _STRUCTURED_REF_RE.sub(" ", _EXTERNAL_REF_RE.sub(" ", formula))


def extract_formula_references(
    formula: str,
    current_sheet: str,
    sheet_lookup: dict[str, str],
) -> set[str]:
    """Extract the ``"Sheet!Cell"`` references a formula depends on.

    Args:
        formula: The formula string, with or without a leading ``=``.
        current_sheet: Sheet the formula lives on — resolves bare refs (``A1``).
        sheet_lookup: Uppercased sheet name → original casing.

    Returns:
        Normalised ``"Sheet!Cell"`` location keys. References to sheets that
        don't exist are dropped: they're a #REF! for Excel to report, not a
        cycle, and guessing at them would only add noise.
    """
    if not formula:
        return set()

    cleaned = _strip_unresolvable_refs(_strip_string_literals(formula))
    refs: set[str] = set()

    for match in _REF_TOKENS_RE.finditer(cleaned):
        sheet_raw, sheet_end_raw, cell_start, cell_end = match.groups()

        if not sheet_raw:
            target_sheets = [current_sheet]
        else:
            start_sheet = sheet_lookup.get(sheet_raw.strip().upper())
            if start_sheet is None:
                continue
            if not sheet_end_raw:
                target_sheets = [start_sheet]
            else:
                # 3D reference (Sheet1:Sheet3!A1) — expand across every sheet
                # between the two names in workbook order, as Excel does.
                end_sheet = sheet_lookup.get(sheet_end_raw.strip().upper())
                if end_sheet is None:
                    continue
                ordered = list(sheet_lookup.values())
                try:
                    i_start, i_end = ordered.index(start_sheet), ordered.index(end_sheet)
                except ValueError:
                    target_sheets = [start_sheet]
                else:
                    if i_start > i_end:
                        i_start, i_end = i_end, i_start
                    target_sheets = ordered[i_start:i_end + 1]

        if cell_end:
            start_norm, end_norm = _normalize_coord(cell_start), _normalize_coord(cell_end)
            if not _is_valid_coord(start_norm) or not _is_valid_coord(end_norm):
                continue
            coords = _expand_range(start_norm, end_norm)
        else:
            coord = _normalize_coord(cell_start)
            if not _is_valid_coord(coord):
                continue
            coords = [coord]

        for sheet in target_sheets:
            for coord in coords:
                refs.add(_format_location(sheet, coord))

    return refs


# ── Cycle detection ──────────────────────────────────────────────────────────


def _build_dependency_graph(
    workbook: Any,
    sheet_names: list[str],
) -> dict[str, set[str]]:
    """Map each formula cell (``"Sheet!Cell"``) to the cells it references."""
    sheet_lookup = _build_sheet_lookup(sheet_names)
    graph: dict[str, set[str]] = {}

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                # `=`-prefixed text is not necessarily a formula: an
                # over-length formula is deliberately stored as an inline
                # string (see helpers._write_formula), and Excel will never
                # evaluate it. Such a cell cannot take part in a cycle, so
                # counting it would only invent one.
                if cell.data_type != "f":
                    continue
                node = _format_location(sheet_name, cell.coordinate)
                graph[node] = extract_formula_references(value, sheet_name, sheet_lookup)

    return graph


def _find_cycles(graph: dict[str, set[str]]) -> set[str]:
    """Return every node that participates in a cycle.

    Standard three-colour DFS: WHITE unvisited, GRAY on the current path, BLACK
    finished. An edge into a GRAY node is a back edge, and everything on the
    path from that node onward is on a cycle.

    The traversal is iterative rather than recursive on purpose — a 500-row
    running-total column is a 500-deep dependency chain, and recursion would
    hit Python's 1000-frame limit on workbooks that are otherwise unremarkable.
    """
    WHITE, GRAY, BLACK = 0, 1, 2
    color: dict[str, int] = dict.fromkeys(graph, WHITE)
    cyclic: set[str] = set()

    for start_node in graph:
        if color[start_node] != WHITE:
            continue

        # `path` mirrors the GRAY nodes in visit order; each stack frame holds
        # the node, an iterator over its remaining dependencies, and the node's
        # index in `path` so the path can be truncated when the node finishes.
        path: list[str] = [start_node]
        stack: list[tuple[str, Any, int]] = [
            (start_node, iter(sorted(graph[start_node])), 0)
        ]
        color[start_node] = GRAY

        while stack:
            node, deps, path_idx = stack[-1]
            descended = False

            for dep in deps:
                if dep not in color:
                    continue        # a literal cell — cannot be part of a cycle
                if color[dep] == GRAY:
                    cyclic.update(path[path.index(dep):])
                    continue        # keep scanning this node's other deps
                if color[dep] == WHITE:
                    color[dep] = GRAY
                    path.append(dep)
                    stack.append((dep, iter(sorted(graph[dep])), len(path) - 1))
                    descended = True
                    break

            if not descended:
                color[node] = BLACK
                del path[path_idx:]
                stack.pop()

    return cyclic


def detect_circular_references(
    xlsx_bytes: bytes,
    sheet_names: list[str],
) -> list[CellError]:
    """Find circular references in a generated workbook.

    Args:
        xlsx_bytes: The saved workbook.
        sheet_names: Sheet names in workbook order (``wb.sheetnames``).

    Returns:
        One :class:`CellError` per cell on a cycle, sorted by location. Empty
        if there are no cycles or the workbook can't be analysed — this is a
        diagnostic, so it never raises and never blocks delivery.
    """
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
    except Exception as e:
        logger.debug("Circular-reference detection skipped (workbook unreadable): %s", e)
        return []

    try:
        graph = _build_dependency_graph(workbook, sheet_names)
    finally:
        workbook.close()

    if not graph:
        return []

    errors = [
        CellError(sheet=sheet, coordinate=coordinate, error_type=CIRCULAR_ERROR_TYPE)
        for sheet, coordinate in (
            _parse_location(location) for location in sorted(_find_cycles(graph))
        )
    ]

    if errors:
        logger.warning(
            "Circular references detected in %d cell(s): %s",
            len(errors),
            format_circular_summary(errors),
        )
    return errors


def format_circular_summary(errors: list[CellError], max_shown: int = 5) -> str:
    """Render detected cycles as a short, human-readable summary line."""
    if not errors:
        return ""
    locations = [err.location for err in errors]
    shown = ", ".join(locations[:max_shown])
    if len(locations) > max_shown:
        shown += f" (and {len(locations) - max_shown} more)"
    return (
        f"{len(locations)} cell(s) on a circular reference: {shown} — a formula "
        f"depends on itself, directly or indirectly. Excel will show a warning "
        f"and resolve these to 0; fix by breaking the cycle."
    )
