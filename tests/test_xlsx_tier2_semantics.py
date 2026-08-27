"""Tests for the Tier 2 Excel behaviour changes.

These pin *intentional* changes to output, as opposed to the pure bug fixes in
test_xlsx_tier1_fixes.py:

- ``B[0]`` is a current-row reference, not a table-relative one.
- Percent and thousands number formats follow the source value's precision.
- Unambiguous digit grouping is recognised in columns with no `types` directive.
"""

import io
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook

from xlsx_tools import markdown_to_excel
from xlsx_tools.helpers import (
    _apply_column_type,
    _normalize_grouped_number,
    _percent_format_for,
    _thousands_format_for,
    adjust_formula_references,
    resolve_cell,
)


def _create_workbook_from_markdown(markdown_content: str) -> Workbook:
    captured = {}

    def fake_upload(file_obj, suffix, **kwargs):
        captured['data'] = file_obj.read()
        file_obj.seek(0)
        return "https://fake-url/test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown_content)

    return load_workbook(io.BytesIO(captured['data']))


class TestRowRelativeReferences:
    """B[n] is an offset from the row the formula lives on."""

    def test_current_row(self):
        assert adjust_formula_references("=B[0]*C[0]", 5, {"T1": 3}, {}) == "=B5*C5"

    def test_moves_with_the_row(self):
        """The bug: every row produced the same formula."""
        rows = [
            adjust_formula_references("=B[0]*C[0]", row, {"T1": 3}, {})
            for row in (4, 5, 6)
        ]
        assert rows == ["=B4*C4", "=B5*C5", "=B6*C6"]

    def test_negative_offset_is_the_row_above(self):
        assert adjust_formula_references("=B[0]-B[-1]", 7, {"T1": 3}, {}) == "=B7-B6"

    def test_positive_offset_is_the_row_below(self):
        assert adjust_formula_references("=B[1]", 7, {"T1": 3}, {}) == "=B8"

    def test_range_is_also_current_row_relative(self):
        assert adjust_formula_references("=SUM(B[-2]:B[0])", 8, {"T1": 3}, {}) == "=SUM(B6:B8)"

    def test_table_relative_form_is_unaffected(self):
        """T1.B[n] still counts data rows from the table's start."""
        assert adjust_formula_references("=T1.B[0]", 9, {"T1": 3}, {}) == "=B4"
        assert adjust_formula_references("=T1.B[2]", 9, {"T1": 3}, {}) == "=B6"

    def test_the_two_forms_are_now_distinct(self):
        """Previously both resolved to B4 — the whole point of the change."""
        result = adjust_formula_references("=T1.B[0]-B[0]", 7, {"T1": 3}, {})
        assert result == "=B4-B7"

    def test_no_table_context_unchanged(self):
        assert adjust_formula_references("=B[0]", 5, {}, {}) == "=B5"

    def test_running_total_end_to_end(self):
        """The canonical case the old behaviour broke."""
        markdown = """| Month | Sales | Cumulative |
|-------|-------|------------|
| Jan   | 100   | =B[0]      |
| Feb   | 200   | =C[-1]+B[0] |
| Mar   | 300   | =C[-1]+B[0] |
"""
        ws = _create_workbook_from_markdown(markdown).active
        assert ws["C2"].value == "=B2"
        assert ws["C3"].value == "=C2+B3"
        assert ws["C4"].value == "=C3+B4"

    def test_cross_sheet_table_form_still_works(self):
        result = adjust_formula_references(
            "=Revenue!T1.B[0]-B[0]", 5, {"T1": 3}, {"Revenue": {"T1": 1}}
        )
        assert result == "=Revenue!B2-B5"


class TestPercentPrecision:
    """Percent formats carry the precision the source text had."""

    @pytest.mark.parametrize("text,expected", [
        ("50", "0%"),
        ("50.5", "0.0%"),
        ("50.25", "0.00%"),
        ("50.125", "0.000%"),
        ("50,5", "0.0%"),           # European decimal comma
        ("50.123456", "0.0000%"),   # capped at 4 decimals
        ("", "0%"),
    ])
    def test_percent_format_for(self, text, expected):
        assert _percent_format_for(text) == expected

    def test_whole_percent_keeps_integer_format(self):
        ws = _create_workbook_from_markdown(
            "| Metric | Rate |\n|--------|------|\n| Growth | 50% |\n"
        ).active
        assert ws["B2"].value == pytest.approx(0.5)
        assert ws["B2"].number_format == "0%"

    def test_fractional_percent_keeps_its_decimal(self):
        """50.5% used to render as '51%' — a number the source never had."""
        ws = _create_workbook_from_markdown(
            "| Metric | Rate |\n|--------|------|\n| Growth | 50.5% |\n"
        ).active
        assert ws["B2"].value == pytest.approx(0.505)
        assert ws["B2"].number_format == "0.0%"

    def test_typed_percent_column_matches(self):
        cell = Workbook().active["A1"]
        _apply_column_type(cell, "45.25%", "percent")
        assert cell.value == pytest.approx(0.4525)
        assert cell.number_format == "0.00%"

    def test_typed_whole_percent_unchanged(self):
        cell = Workbook().active["A1"]
        _apply_column_type(cell, "45%", "percent")
        assert cell.number_format == "0%"


class TestThousandsFormat:
    """Grouped formats no longer round decimals away."""

    @pytest.mark.parametrize("value,expected", [
        (1500, "#,##0"),
        (1500.75, "#,##0.00"),
        (1000, "#,##0"),
        (-2500.5, "#,##0.00"),
    ])
    def test_thousands_format_for(self, value, expected):
        assert _thousands_format_for(value) == expected

    def test_decimal_thousands_not_rounded(self):
        """1500.75 used to display as '1,501'."""
        ws = _create_workbook_from_markdown(
            "| Item | Amount |\n|------|--------|\n| A | 1500.75 |\n"
        ).active
        assert ws["B2"].value == pytest.approx(1500.75)
        assert ws["B2"].number_format == "#,##0.00"

    def test_whole_thousands_unchanged(self):
        ws = _create_workbook_from_markdown(
            "| Item | Amount |\n|------|--------|\n| A | 1500 |\n"
        ).active
        assert ws["B2"].number_format == "#,##0"

    def test_negative_thousands_now_grouped(self):
        """`>= 1000` left every negative value unformatted."""
        ws = _create_workbook_from_markdown(
            "| Item | Amount |\n|------|--------|\n| A | -5000 |\n"
        ).active
        assert ws["B2"].number_format == "#,##0"

    @pytest.mark.parametrize("raw", ["0.001", "0.5", "12.5", "999.99"])
    def test_small_values_keep_general_format(self, raw):
        """Below 1000 nothing is imposed, so precision is never hidden."""
        ws = _create_workbook_from_markdown(
            f"| Item | Amount |\n|------|--------|\n| A | {raw} |\n"
        ).active
        assert ws["B2"].value == pytest.approx(float(raw))
        assert ws["B2"].number_format == "General"

    def test_typed_number_column_matches(self):
        cell = Workbook().active["A1"]
        _apply_column_type(cell, "1500.75", "number")
        assert cell.number_format == "#,##0.00"


class TestUnambiguousDigitGrouping:
    """Untyped columns accept grouping only where the reading is certain."""

    @pytest.mark.parametrize("raw,expected", [
        ("1,234", "1234"),
        ("1,234.56", "1234.56"),
        ("1,234,567.89", "1234567.89"),
        ("1.234,56", "1234.56"),
        ("1.234.567", "1234567"),
    ])
    def test_unambiguous_grouping_normalised(self, raw, expected):
        assert _normalize_grouped_number(raw) == expected

    @pytest.mark.parametrize("raw", [
        "1,5",        # European decimal — 1.5, not 15
        "1,23",
        "1,2345",
        "1.234",      # plain English decimal — must stay 1.234
        "12.5",
        "Smith, J",
    ])
    def test_ambiguous_or_non_numeric_left_alone(self, raw):
        assert _normalize_grouped_number(raw) == raw

    def test_grouped_value_becomes_a_number(self):
        assert resolve_cell("1,234").value == pytest.approx(1234.0)

    def test_ambiguous_comma_stays_text(self):
        """Auto-detection must not guess; `types: number` is how you opt in."""
        assert resolve_cell("1,5").value == "1,5"

    def test_bare_dot_decimal_unchanged(self):
        assert resolve_cell("1.234").value == pytest.approx(1.234)

    def test_text_with_comma_unchanged(self):
        assert resolve_cell("Smith, J").value == "Smith, J"

    def test_typed_column_still_resolves_ambiguous_forms(self):
        """The explicit directive opts into the locale heuristic."""
        cell = Workbook().active["A1"]
        _apply_column_type(cell, "1,5", "number")
        assert cell.value == pytest.approx(1.5)

    def test_grouped_value_end_to_end(self):
        ws = _create_workbook_from_markdown(
            "| Item | Amount |\n|------|--------|\n| A | 1,234 |\n"
        ).active
        assert ws["B2"].value == pytest.approx(1234.0)
        assert ws["B2"].number_format == "#,##0"
