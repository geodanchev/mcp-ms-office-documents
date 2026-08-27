"""Regression tests for the Tier 1 Excel correctness fixes.

Each test here pins a behaviour that was demonstrably wrong before, so the
fixes can't silently regress. Grouped by the bug they cover.
"""

import io
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook

from xlsx_tools import markdown_to_excel
from xlsx_tools.circular_refs import (
    CIRCULAR_ERROR_TYPE,
    detect_circular_references,
    extract_formula_references,
)
from xlsx_tools.helpers import (
    _apply_column_type,
    _number_format_for_type,
    _parse_types_directive,
    _strip_thousands_separators,
    adjust_formula_references,
)


def _create_workbook_from_markdown(markdown_content: str) -> Workbook:
    """Run markdown_to_excel, intercepting the bytes before upload."""
    captured = {}

    def fake_upload(file_obj, suffix, **kwargs):
        captured['data'] = file_obj.read()
        file_obj.seek(0)
        return "https://fake-url/test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown_content)

    return load_workbook(io.BytesIO(captured['data']))


def _cell(type_spec: str, raw_text: str):
    """Apply a column type to a scratch cell and return it."""
    cell = Workbook().active["A1"]
    _apply_column_type(cell, raw_text, type_spec)
    return cell


class TestCrossSheetRangePrefix:
    """The sheet prefix belongs on the first range endpoint only."""

    def test_cross_sheet_sum_emits_single_prefix(self):
        result = adjust_formula_references(
            "=Sales!T1.SUM(B[0]:D[0])", 10, {}, {"Sales": {"T1": 3}}
        )
        assert result == "=SUM(Sales!B4:D4)"

    def test_quoted_sheet_name_also_single_prefix(self):
        result = adjust_formula_references(
            "=Sales Data!T1.SUM(B[0]:D[0])", 10, {}, {"Sales Data": {"T1": 3}}
        )
        assert result == "=SUM('Sales Data'!B4:D4)"

    def test_cross_sheet_range_unchanged(self):
        """The non-function range form was already correct — keep it that way."""
        result = adjust_formula_references(
            "=SUM(Data!T1.B[0]:T1.B[2])", 10, {}, {"Data": {"T1": 1}}
        )
        assert result == "=SUM(Data!B2:B4)"


class TestTypesDirectiveCommaParsing:
    """A comma inside a number format must not be read as a column boundary."""

    def test_literal_format_with_comma_survives(self):
        assert _parse_types_directive("text, number:#,##0.00, percent") == [
            "text", "number:#,##0.00", "percent",
        ]

    def test_column_count_is_preserved(self):
        """The bug shifted every later column by one — pin the count."""
        specs = _parse_types_directive("text, number:#,##0, currency:$, bool")
        assert len(specs) == 4
        assert specs[2] == "currency:$"

    def test_unspecified_column_between_formats(self):
        assert _parse_types_directive("number:#,##0, , text") == [
            "number:#,##0", None, "text",
        ]

    def test_multi_section_format_with_commas(self):
        assert _parse_types_directive("text, number:#,##0;[Red](#,##0), bool") == [
            "text", "number:#,##0;[Red](#,##0)", "bool",
        ]

    def test_plain_directive_unchanged(self):
        assert _parse_types_directive("text, currency:$, date, bool") == [
            "text", "currency:$", "date", "bool",
        ]

    def test_empty_value(self):
        assert _parse_types_directive("") == []

    def test_number_format_lands_on_the_cell(self):
        """End-to-end: the format reaches the cell instead of being truncated."""
        markdown = """<!-- types: text, number:#,##0.00 -->
| Item | Amount |
|------|--------|
| A    | 1234.5 |
"""
        ws = _create_workbook_from_markdown(markdown).active
        assert ws.cell(row=2, column=2).value == pytest.approx(1234.5)
        assert ws.cell(row=2, column=2).number_format == "#,##0.00"


class TestThousandsSeparators:
    """European decimal commas must not be read as thousands separators."""

    @pytest.mark.parametrize("raw,expected", [
        ("1,5", "1.5"),          # European decimal
        ("1,234", "1234"),       # thousands
        ("1,234.56", "1234.56"), # English
        ("1.234,56", "1234.56"), # European
        ("1234", "1234"),        # untouched
        ("12.5", "12.5"),        # untouched
        ("abc", "abc"),          # non-numeric passes through
    ])
    def test_strip_thousands_separators(self, raw, expected):
        assert _strip_thousands_separators(raw) == expected

    def test_number_type_european_decimal(self):
        """Was 15.0 — a silent 10x error."""
        assert _cell("number", "1,5").value == pytest.approx(1.5)

    def test_number_type_thousands(self):
        assert _cell("number", "1,234").value == pytest.approx(1234.0)

    def test_currency_type_european_decimal(self):
        assert _cell("currency:€", "1,5 €").value == pytest.approx(1.5)

    def test_currency_type_english_thousands(self):
        assert _cell("currency:$", "$1,234.56").value == pytest.approx(1234.56)


class TestAccountingNegatives:
    """Parenthesised currency amounts are negative numbers, not text."""

    @pytest.mark.parametrize("raw,expected", [
        ("($1,234)", -1234.0),
        ("($1,234.56)", -1234.56),
        ("(1,234)", -1234.0),
    ])
    def test_parenthesised_amount_is_negative(self, raw, expected):
        cell = _cell("currency:$", raw)
        assert cell.value == pytest.approx(expected)
        assert cell.number_format == "$#,##0.00"

    def test_positive_amount_unaffected(self):
        assert _cell("currency:$", "$1,234").value == pytest.approx(1234.0)

    def test_unparseable_still_falls_back_to_text(self):
        assert _cell("currency:$", "(n/a)").value == "(n/a)"


class TestFormulasInTypedColumns:
    """A formula in a typed column must be resolved, not coerced."""

    def test_formula_is_resolved_not_stringified(self):
        markdown = """<!-- types: text, currency:$ -->
| Item  | Amount |
|-------|--------|
| A     | 100    |
| B     | 200    |
| Total | =SUM(T1.B[0]:T1.B[1]) |
"""
        ws = _create_workbook_from_markdown(markdown).active
        total = ws.cell(row=4, column=2)
        assert total.value == "=SUM(B2:B3)", "table refs must still be resolved"
        assert total.number_format == "$#,##0.00", "column format must be applied"

    def test_formula_keeps_percent_format(self):
        markdown = """<!-- types: text, percent -->
| Metric | Rate |
|--------|------|
| A      | 50%  |
| Calc   | =B2*2 |
"""
        ws = _create_workbook_from_markdown(markdown).active
        assert ws.cell(row=3, column=2).value == "=B2*2"
        assert ws.cell(row=3, column=2).number_format == "0%"

    def test_literal_values_in_typed_column_still_coerced(self):
        """The formula carve-out must not disable coercion for normal cells."""
        markdown = """<!-- types: text, currency:$ -->
| Item | Amount |
|------|--------|
| A    | $1,500 |
"""
        ws = _create_workbook_from_markdown(markdown).active
        assert ws.cell(row=2, column=2).value == pytest.approx(1500.0)

    @pytest.mark.parametrize("spec,expected", [
        ("currency:$", "$#,##0.00"),
        ("currency:€", "#,##0.00 €"),
        ("currency:XYZ", '#,##0.00 "XYZ"'),
        ("percent", "0%"),
        ("number:0.000", "0.000"),
        ("date:DD.MM.YYYY", "DD.MM.YYYY"),
        ("number", None),      # bare form is value-derived
        ("date", None),        # bare form is value-derived
        ("text", None),
        ("bool", None),
        (None, None),
    ])
    def test_number_format_for_type(self, spec, expected):
        assert _number_format_for_type(spec) == expected


class TestUnresolvedReferenceWarnings:
    """Bad references still resolve, but must no longer be silent."""

    def test_unknown_table_warns(self, caplog):
        with caplog.at_level("WARNING"):
            adjust_formula_references("=T9.B[0]", 5, {"T1": 3}, {})
        assert "T9" in caplog.text

    def test_unknown_sheet_warns(self, caplog):
        with caplog.at_level("WARNING"):
            adjust_formula_references(
                "=Revenue!T1.B[0]", 5, {}, {"Revenue Model": {"T1": 1}}
            )
        assert "Revenue" in caplog.text and "#REF!" in caplog.text

    def test_known_sheet_does_not_warn(self, caplog):
        with caplog.at_level("WARNING"):
            adjust_formula_references("=Revenue!T1.B[0]", 5, {}, {"Revenue": {"T1": 1}})
        assert caplog.text == ""

    def test_colliding_sheet_names_warn(self, caplog):
        markdown = """## Sheet: Report
| A |
|---|
| 1 |

## Sheet: Report
| B |
|---|
| 2 |
"""
        with caplog.at_level("WARNING"):
            wb = _create_workbook_from_markdown(markdown)
        assert "collides" in caplog.text
        assert len(wb.sheetnames) == 2


class TestFormulaReferenceExtraction:
    """Reference parsing underpinning cycle detection."""

    def test_local_and_range_refs(self):
        refs = extract_formula_references("=A1+SUM(B1:B3)", "S", {"S": "S"})
        assert refs == {"S!A1", "S!B1", "S!B2", "S!B3"}

    def test_absolute_refs_normalised(self):
        assert extract_formula_references("=$A$1", "S", {"S": "S"}) == {"S!A1"}

    def test_cross_sheet_ref(self):
        refs = extract_formula_references(
            "=Data!A1", "S", {"S": "S", "DATA": "Data"}
        )
        assert refs == {"Data!A1"}

    def test_unknown_sheet_dropped(self):
        assert extract_formula_references("=Ghost!A1", "S", {"S": "S"}) == set()

    def test_string_literals_are_not_references(self):
        """='see cell A1' must not produce an A1 edge."""
        assert extract_formula_references('="see cell A1"', "S", {"S": "S"}) == set()

    def test_escaped_quotes_inside_literal(self):
        assert extract_formula_references(
            '=IF(A1>0,"a""B2""c","")', "S", {"S": "S"}
        ) == {"S!A1"}

    def test_structured_reference_not_mistaken_for_coord(self):
        assert "S!BLE1" not in extract_formula_references(
            "=Table1[Revenue]", "S", {"S": "S"}
        )

    def test_huge_range_falls_back_to_corners(self):
        refs = extract_formula_references("=SUM(A1:A100000)", "S", {"S": "S"})
        assert refs == {"S!A1", "S!A100000"}


class TestCircularReferenceDetection:
    def _detect(self, cells: dict[str, str], sheet: str = "Sheet"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        for coord, value in cells.items():
            ws[coord] = value
        buf = io.BytesIO()
        wb.save(buf)
        return detect_circular_references(buf.getvalue(), wb.sheetnames)

    def test_self_reference(self):
        errors = self._detect({"A1": "=A1+1"})
        assert [e.location for e in errors] == ["Sheet!A1"]
        assert errors[0].error_type == CIRCULAR_ERROR_TYPE

    def test_two_cell_cycle(self):
        errors = self._detect({"A1": "=B1+1", "B1": "=A1+1"})
        assert {e.location for e in errors} == {"Sheet!A1", "Sheet!B1"}

    def test_three_cell_cycle(self):
        errors = self._detect({"A1": "=B1", "B1": "=C1", "C1": "=A1"})
        assert {e.location for e in errors} == {"Sheet!A1", "Sheet!B1", "Sheet!C1"}

    def test_cycle_through_a_range(self):
        errors = self._detect({"A1": "1", "A2": "2", "A3": "=SUM(A1:A3)"})
        assert [e.location for e in errors] == ["Sheet!A3"]

    def test_acyclic_chain_is_clean(self):
        assert self._detect({"A1": "1", "A2": "=A1+1", "A3": "=A2+1"}) == []

    def test_diamond_dependency_is_not_a_cycle(self):
        """A1 → B1,C1 → D1: shared descendants must not be reported."""
        assert self._detect(
            {"D1": "1", "B1": "=D1", "C1": "=D1", "A1": "=B1+C1"}
        ) == []

    def test_self_referential_label_is_not_a_cycle(self):
        """=A1&" totals" placed in B1 references A1, not itself."""
        assert self._detect({"A1": "x", "B1": '=A1&" totals"'}) == []

    def test_no_formulas_is_clean(self):
        assert self._detect({"A1": "x", "A2": "1"}) == []

    def test_long_chain_does_not_hit_recursion_limit(self):
        """A 1500-deep chain would blow a recursive DFS's stack."""
        cells = {"A1": "1"}
        for row in range(2, 1502):
            cells[f"A{row}"] = f"=A{row - 1}+1"
        assert self._detect(cells) == []

    def test_long_chain_closed_into_a_cycle_is_found(self):
        cells = {"A1": "=A1500"}
        for row in range(2, 1501):
            cells[f"A{row}"] = f"=A{row - 1}+1"
        errors = self._detect(cells)
        assert len(errors) == 1500

    def test_cross_sheet_cycle(self):
        wb = Workbook()
        first = wb.active
        first.title = "One"
        second = wb.create_sheet("Two")
        first["A1"] = "=Two!A1"
        second["A1"] = "=One!A1"
        buf = io.BytesIO()
        wb.save(buf)
        errors = detect_circular_references(buf.getvalue(), wb.sheetnames)
        assert {e.location for e in errors} == {"One!A1", "Two!A1"}

    def test_quoted_sheet_name_round_trips(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "My Sheet"
        ws["A1"] = "=A1+1"
        buf = io.BytesIO()
        wb.save(buf)
        errors = detect_circular_references(buf.getvalue(), wb.sheetnames)
        assert errors[0].location == "'My Sheet'!A1"
        assert errors[0].sheet == "My Sheet"
        assert errors[0].coordinate == "A1"

    def test_unreadable_bytes_return_empty(self):
        assert detect_circular_references(b"not an xlsx", ["Sheet"]) == []

    def test_detection_runs_during_generation(self, caplog):
        markdown = """| Item | Value |
|------|-------|
| A    | =B2+1 |
"""
        with caplog.at_level("WARNING"):
            _create_workbook_from_markdown(markdown)
        assert "circular" in caplog.text.lower()

    def test_clean_workbook_logs_nothing(self, caplog):
        markdown = """| Item | Value |
|------|-------|
| A    | 1     |
| B    | =B2*2 |
"""
        with caplog.at_level("WARNING"):
            _create_workbook_from_markdown(markdown)
        assert "circular" not in caplog.text.lower()
