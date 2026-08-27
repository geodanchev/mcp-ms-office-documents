"""Tests for the `styles:` directive and template-defined named styles."""

import io
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

from xlsx_tools import markdown_to_excel
from xlsx_tools.styles import (
    MAX_STYLED_CELLS,
    StyleSpec,
    TemplateStyles,
    apply_style_spec,
    load_template_styles,
    parse_color,
    parse_style_spec,
    parse_styles_directive,
)


def _workbook(markdown_content: str) -> Workbook:
    captured = {}

    def fake_upload(file_obj, suffix, **kwargs):
        captured['data'] = file_obj.read()
        file_obj.seek(0)
        return "https://fake-url/test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown_content)
    return load_workbook(io.BytesIO(captured['data']))


def _rgb(color) -> str | None:
    """openpyxl reports colours as 8-digit ARGB; compare on the RGB tail."""
    value = getattr(color, "rgb", None)
    return value[-6:] if isinstance(value, str) else None


class TestColorParsing:
    @pytest.mark.parametrize("raw,expected", [
        ("FFFF00", "FFFF00"),
        ("#FFFF00", "FFFF00"),
        ("ffff00", "FFFF00"),
        ("FFFFFF00", "FFFFFF00"),
        ("yellow", "FFFF00"),
        ("Red", "FF0000"),
        ("  blue  ", "0070C0"),
    ])
    def test_valid_colors(self, raw, expected):
        assert parse_color(raw) == expected

    def test_us_and_uk_grey_both_work(self):
        assert parse_color("grey") == parse_color("gray")
        assert parse_color("lightgrey") == parse_color("lightgray")

    @pytest.mark.parametrize("raw", ["", "notacolor", "FFF", "GGGGGG", "#12345"])
    def test_invalid_colors(self, raw):
        assert parse_color(raw) is None


class TestStyleSpecParsing:
    def test_background(self):
        assert parse_style_spec("bg:yellow").background == "FFFF00"

    def test_font_color(self):
        assert parse_style_spec("color:red").font_color == "FF0000"

    @pytest.mark.parametrize("flag", ["bold", "italic", "underline"])
    def test_flags(self, flag):
        assert getattr(parse_style_spec(flag), flag) is True

    def test_combined(self):
        spec = parse_style_spec("bg:yellow;color:red;bold;italic")
        assert (spec.background, spec.font_color, spec.bold, spec.italic) == \
               ("FFFF00", "FF0000", True, True)

    def test_named_style(self):
        assert parse_style_spec("style:Total").named_style == "Total"

    def test_aliases(self):
        assert parse_style_spec("background:yellow").background == "FFFF00"
        assert parse_style_spec("fill:yellow").background == "FFFF00"
        assert parse_style_spec("fg:red").font_color == "FF0000"

    def test_unknown_attribute_is_skipped_not_fatal(self, caplog):
        with caplog.at_level("WARNING"):
            spec = parse_style_spec("bogus;bold")
        assert spec.bold is True
        assert "bogus" in caplog.text

    def test_bad_color_warns_and_is_dropped(self, caplog):
        with caplog.at_level("WARNING"):
            spec = parse_style_spec("bg:chartreuse;bold")
        assert spec.background is None
        assert spec.bold is True
        assert "chartreuse" in caplog.text

    def test_empty_spec(self):
        assert parse_style_spec("").is_empty()


class TestDirectiveTargets:
    def test_single_absolute_cell(self):
        assert set(parse_styles_directive("B2=bg:yellow")) == {"B2"}

    def test_multiple_entries(self):
        assert set(parse_styles_directive("B2=bg:yellow, C3=bold")) == {"B2", "C3"}

    def test_range_expands_to_full_rectangle(self):
        """Corners-only would silently miss the interior of a block."""
        assert set(parse_styles_directive("B2:C3=bold")) == {"B2", "B3", "C2", "C3"}

    def test_single_column_range(self):
        assert set(parse_styles_directive("B2:B4=bold")) == {"B2", "B3", "B4"}

    def test_reversed_range_is_normalised(self):
        assert set(parse_styles_directive("C3:B2=bold")) == {"B2", "B3", "C2", "C3"}

    def test_table_relative_rows(self):
        """B[0] is the table's first data row; the header sits at start_row."""
        styles = parse_styles_directive("B[0]=bold, B[2]=italic", table_start_row=5)
        assert set(styles) == {"B6", "B8"}

    def test_table_relative_header_row(self):
        assert set(parse_styles_directive("B[-1]=bold", table_start_row=5)) == {"B5"}

    def test_table_relative_range(self):
        assert set(parse_styles_directive("B[0]:B[2]=bold", table_start_row=5)) == \
               {"B6", "B7", "B8"}

    def test_relative_target_needs_an_anchor(self, caplog):
        with caplog.at_level("WARNING"):
            assert parse_styles_directive("B[0]=bold", table_start_row=None) == {}

    def test_later_entry_wins_on_overlap(self):
        styles = parse_styles_directive("B2:B4=bg:yellow, B3=bg:red")
        assert styles["B2"].background == "FFFF00"
        assert styles["B3"].background == "FF0000"

    @pytest.mark.parametrize("entry", ["B2", "=bold", "ZZZZ9=bold", "B0=bold", "2B=bold"])
    def test_malformed_entries_skipped(self, entry, caplog):
        with caplog.at_level("WARNING"):
            assert parse_styles_directive(entry) == {}

    def test_oversized_range_is_dropped_not_truncated(self, caplog):
        with caplog.at_level("WARNING"):
            result = parse_styles_directive("A1:Z100000=bold")
        assert result == {}
        assert str(MAX_STYLED_CELLS) in caplog.text

    def test_empty_directive(self):
        assert parse_styles_directive("") == {}


class TestApplyStyleSpec:
    def test_background_applied(self):
        cell = Workbook().active["A1"]
        apply_style_spec(cell, StyleSpec(background="FFFF00"))
        assert _rgb(cell.fill.start_color) == "FFFF00"

    def test_font_color_applied(self):
        cell = Workbook().active["A1"]
        apply_style_spec(cell, StyleSpec(font_color="FF0000"))
        assert _rgb(cell.font.color) == "FF0000"

    def test_flags_applied(self):
        cell = Workbook().active["A1"]
        apply_style_spec(cell, StyleSpec(bold=True, italic=True, underline=True))
        assert cell.font.bold and cell.font.italic
        assert cell.font.underline == "single"

    def test_existing_font_family_and_size_preserved(self):
        cell = Workbook().active["A1"]
        cell.font = Font(name="Arial", size=14)
        apply_style_spec(cell, StyleSpec(background="FFFF00", bold=True))
        assert cell.font.name == "Arial"
        assert cell.font.size == 14

    def test_unset_flags_do_not_clear_existing(self):
        """bg:yellow on a bold header must not un-bold it."""
        cell = Workbook().active["A1"]
        cell.font = Font(bold=True)
        apply_style_spec(cell, StyleSpec(font_color="FF0000"))
        assert cell.font.bold is True

    def test_unknown_named_style_warns_and_is_ignored(self, caplog):
        cell = Workbook().active["A1"]
        with caplog.at_level("WARNING"):
            apply_style_spec(cell, StyleSpec(named_style="Nope"), available_styles=set())
        assert "Nope" in caplog.text

    def test_inline_attributes_layer_over_named_style(self):
        wb = Workbook()
        ns = NamedStyle(name="Total")
        ns.font = Font(bold=True, color="0000FF")
        ns.fill = PatternFill("solid", start_color="DDDDDD")
        wb.add_named_style(ns)
        cell = wb.active["A1"]

        apply_style_spec(cell, StyleSpec(named_style="Total", background="FFFF00"),
                         available_styles={"Total"})
        assert _rgb(cell.fill.start_color) == "FFFF00", "inline bg must win"
        assert cell.font.bold is True, "named style's font must survive"


class TestEndToEnd:
    def test_background_reaches_the_cell(self):
        ws = _workbook(
            "<!-- styles: B2=bg:yellow -->\n"
            "| Item | Value |\n|------|-------|\n| A | 1 |\n"
        ).active
        assert _rgb(ws["B2"].fill.start_color) == "FFFF00"

    def test_font_color_and_bold(self):
        ws = _workbook(
            "<!-- styles: A2=color:red;bold -->\n"
            "| Item | Value |\n|------|-------|\n| A | 1 |\n"
        ).active
        assert _rgb(ws["A2"].font.color) == "FF0000"
        assert ws["A2"].font.bold is True

    def test_table_relative_targets_end_to_end(self):
        """B[0] must land on the first data row even below a heading."""
        ws = _workbook(
            "# Report\n\n"
            "<!-- styles: B[0]=bg:yellow -->\n"
            "| Item | Value |\n|------|-------|\n| A | 1 |\n| B | 2 |\n"
        ).active
        highlighted = [
            c.coordinate for row in ws.iter_rows() for c in row
            if _rgb(c.fill.start_color) == "FFFF00"
        ]
        assert highlighted == ["B4"]

    def test_styles_override_the_formula_fill(self):
        """An explicit instruction must beat the renderer's own styling."""
        ws = _workbook(
            "<!-- styles: B3=bg:yellow -->\n"
            "| Item | Value |\n|------|-------|\n| A | 1 |\n| T | =SUM(T1.B[0]:T1.B[0]) |\n"
        ).active
        assert _rgb(ws["B3"].fill.start_color) == "FFFF00"

    def test_styles_override_the_header_fill(self):
        ws = _workbook(
            "<!-- styles: A1=bg:red -->\n"
            "| Item | Value |\n|------|-------|\n| A | 1 |\n"
        ).active
        assert _rgb(ws["A1"].fill.start_color) == "FF0000"

    def test_number_format_survives_styling(self):
        """Styling sets colours, not data — the type work must be untouched."""
        ws = _workbook(
            "<!-- types: text, currency:$ -->\n"
            "<!-- styles: B2=bg:yellow -->\n"
            "| Item | Value |\n|------|-------|\n| A | $1,500 |\n"
        ).active
        assert ws["B2"].value == pytest.approx(1500.0)
        assert ws["B2"].number_format == "$#,##0.00"
        assert _rgb(ws["B2"].fill.start_color) == "FFFF00"

    def test_range_styling(self):
        ws = _workbook(
            "<!-- styles: A2:B3=bg:lightgrey -->\n"
            "| Item | Value |\n|------|-------|\n| A | 1 |\n| B | 2 |\n"
        ).active
        for coord in ("A2", "B2", "A3", "B3"):
            assert _rgb(ws[coord].fill.start_color) == "D9D9D9", coord

    def test_no_directive_leaves_output_unchanged(self):
        plain = _workbook("| Item | Value |\n|------|-------|\n| A | 1 |\n").active
        assert plain["B2"].fill.fill_type in (None, "none")

    def test_bad_directive_does_not_fail_generation(self, caplog):
        with caplog.at_level("WARNING"):
            ws = _workbook(
                "<!-- styles: nonsense -->\n"
                "| Item | Value |\n|------|-------|\n| A | 1 |\n"
            ).active
        assert ws["B2"].value == 1

    def test_directive_applies_to_one_table_only(self):
        ws = _workbook(
            "<!-- styles: B[0]=bg:yellow -->\n"
            "| A | B |\n|---|---|\n| 1 | 2 |\n\n"
            "| C | D |\n|---|---|\n| 3 | 4 |\n"
        ).active
        highlighted = [
            c.coordinate for row in ws.iter_rows() for c in row
            if _rgb(c.fill.start_color) == "FFFF00"
        ]
        assert highlighted == ["B2"]


class TestTemplateNamedStyles:
    @staticmethod
    def _template_bytes() -> bytes:
        wb = Workbook()
        for name, color in (("Total", "FFFF00"), ("Warning", "FF0000")):
            ns = NamedStyle(name=name)
            ns.font = Font(bold=True, color="FFFFFF")
            ns.fill = PatternFill("solid", start_color=color)
            wb.add_named_style(ns)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _patched_template(self, tmp_path):
        path = tmp_path / "custom_xlsx_template.xlsx"
        path.write_bytes(self._template_bytes())
        return patch("template_utils.find_file_in_template_dirs", return_value=path)

    def test_no_template_is_not_an_error(self):
        with patch("template_utils.find_file_in_template_dirs", return_value=None):
            loaded = load_template_styles()
        assert loaded.names == set()

    def test_unreadable_template_is_not_fatal(self, tmp_path):
        bad = tmp_path / "custom_xlsx_template.xlsx"
        bad.write_bytes(b"not an xlsx")
        with patch("template_utils.find_file_in_template_dirs", return_value=bad):
            assert load_template_styles().names == set()

    def test_styles_are_loaded(self, tmp_path):
        with self._patched_template(tmp_path):
            loaded = load_template_styles()
        assert loaded.names == {"Total", "Warning"}

    def test_builtin_normal_style_is_not_offered(self, tmp_path):
        with self._patched_template(tmp_path):
            assert "Normal" not in load_template_styles().names

    def test_registering_is_idempotent(self, tmp_path):
        with self._patched_template(tmp_path):
            loaded = load_template_styles()
        wb = Workbook()
        assert loaded.register_into(wb) == {"Total", "Warning"}
        assert loaded.register_into(wb) == {"Total", "Warning"}

    def test_named_style_applied_end_to_end(self, tmp_path):
        with self._patched_template(tmp_path):
            ws = _workbook(
                "<!-- styles: B2=style:Total -->\n"
                "| Item | Value |\n|------|-------|\n| A | 1 |\n"
            ).active
        assert _rgb(ws["B2"].fill.start_color) == "FFFF00"
        assert ws["B2"].font.bold is True

    def test_unknown_style_name_warns_but_ships(self, tmp_path, caplog):
        with self._patched_template(tmp_path), caplog.at_level("WARNING"):
            ws = _workbook(
                "<!-- styles: B2=style:Missing -->\n"
                "| Item | Value |\n|------|-------|\n| A | 1 |\n"
            ).active
        assert "Missing" in caplog.text
        assert ws["B2"].value == 1

    def test_empty_template_styles_register_cleanly(self):
        assert TemplateStyles().register_into(Workbook()) == set()


class TestNamedStyleDoesNotClobberCellFormat:
    """A NamedStyle bundles number format, border and alignment with the font.

    `cell.style = name` replaces all five at once, so a style that only
    declares a font would otherwise reset the other three to their defaults —
    silently undoing the table renderer's work.
    """

    @staticmethod
    def _template_with(styles: dict) -> bytes:
        wb = Workbook()
        for name, configure in styles.items():
            ns = NamedStyle(name=name)
            ns.font = Font(bold=True, color="FFFFFF")
            configure(ns)
            wb.add_named_style(ns)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _with_template(self, tmp_path, styles):
        path = tmp_path / "custom_xlsx_template.xlsx"
        path.write_bytes(self._template_with(styles))
        return patch("template_utils.find_file_in_template_dirs", return_value=path)

    def test_font_only_style_preserves_number_format(self, tmp_path):
        """A totals row given style:Total in a currency column kept $#,##0.00."""
        with self._with_template(tmp_path, {"Total": lambda ns: None}):
            ws = _workbook(
                "<!-- types: text, currency:$ -->\n"
                "<!-- styles: B2=style:Total -->\n"
                "| Item | Value |\n|------|-------|\n| A | $1,500 |\n"
            ).active
        assert ws["B2"].value == pytest.approx(1500.0)
        assert ws["B2"].number_format == "$#,##0.00"
        assert ws["B2"].font.bold is True, "the style's font must still apply"

    def test_font_only_style_preserves_border_and_alignment(self, tmp_path):
        with self._with_template(tmp_path, {"Total": lambda ns: None}):
            ws = _workbook(
                "<!-- styles: B2=style:Total -->\n"
                "| Item | Value |\n|------|-------|\n| A | 1 |\n"
            ).active
        assert ws["B2"].border.left.style == "thin", "table border must survive"
        assert ws["B2"].alignment.horizontal == "right", "alignment must survive"

    def test_style_that_declares_a_format_still_applies_it(self, tmp_path):
        """The other direction: a deliberate format in the style must win.

        A blanket restore would discard it.
        """
        def set_percent(ns):
            ns.number_format = "0.00%"

        with self._with_template(tmp_path, {"Pct": set_percent}):
            ws = _workbook(
                "<!-- types: text, currency:$ -->\n"
                "<!-- styles: B2=style:Pct -->\n"
                "| Item | Value |\n|------|-------|\n| A | $1,500 |\n"
            ).active
        assert ws["B2"].number_format == "0.00%"

    def test_style_that_declares_alignment_still_applies_it(self, tmp_path):
        def set_center(ns):
            ns.alignment = Alignment(horizontal="center")

        with self._with_template(tmp_path, {"Mid": set_center}):
            ws = _workbook(
                "<!-- styles: B2=style:Mid -->\n"
                "| Item | Value |\n|------|-------|\n| A | 1 |\n"
            ).active
        assert ws["B2"].alignment.horizontal == "center"

    def test_style_combined_with_inline_attributes(self, tmp_path):
        with self._with_template(tmp_path, {"Total": lambda ns: None}):
            ws = _workbook(
                "<!-- types: text, currency:$ -->\n"
                "<!-- styles: B2=style:Total;bg:yellow -->\n"
                "| Item | Value |\n|------|-------|\n| A | $1,500 |\n"
            ).active
        assert _rgb(ws["B2"].fill.start_color) == "FFFF00", "inline bg wins"
        assert ws["B2"].number_format == "$#,##0.00", "format still preserved"
        assert ws["B2"].font.bold is True
