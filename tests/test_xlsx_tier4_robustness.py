"""Tests for the Tier 4 Excel robustness fixes.

Covers the cases that previously produced a workbook Excel refuses to open, a
formula Excel cannot parse, or a silently blank cell — plus the collapse of the
two duplicated workbook-building code paths.
"""

import io
import re
import zipfile
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook

from xlsx_tools import _markdown_to_excel_buffer, markdown_to_excel
from xlsx_tools.base_xlsx_tool import _build_workbook
from xlsx_tools.helpers import (
    MAX_FORMULA_LENGTH,
    _quote_sheet_name,
    _unquote_sheet_name,
    adjust_formula_references,
    apply_cell_formatting,
    resolve_cell,
)


def _workbook_bytes(markdown_content: str) -> bytes:
    captured = {}

    def fake_upload(file_obj, suffix, **kwargs):
        captured['data'] = file_obj.read()
        file_obj.seek(0)
        return "https://fake-url/test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown_content)
    return captured['data']


def _table_column_names(xlsx_bytes: bytes) -> list[str]:
    """Read the tableColumn names Excel will validate on open."""
    zf = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    table_parts = [n for n in zf.namelist() if n.startswith("xl/tables/")]
    assert table_parts, "expected an Excel Table part"
    xml = zf.read(table_parts[0]).decode()
    return re.findall(r'<tableColumn[^>]*\bname="([^"]*)"', xml)


class TestSheetNameQuoting:
    """Sheet names that aren't plain identifiers must round-trip."""

    @pytest.mark.parametrize("name,expected", [
        ("Revenue", "Revenue"),
        ("Sales Data", "'Sales Data'"),
        ("P&L", "'P&L'"),
        ("John's Data", "'John''s Data'"),
        ("FY-2024", "'FY-2024'"),
    ])
    def test_quote_sheet_name(self, name, expected):
        assert _quote_sheet_name(name) == expected

    @pytest.mark.parametrize("raw,expected", [
        ("Revenue", "Revenue"),
        ("'Sales Data'", "Sales Data"),
        ("'John''s Data'", "John's Data"),
        ("  'P&L'  ", "P&L"),
    ])
    def test_unquote_sheet_name(self, raw, expected):
        assert _unquote_sheet_name(raw) == expected

    def test_quote_unquote_round_trip(self):
        for name in ["Revenue", "Sales Data", "P&L", "John's Data", "A'B'C"]:
            assert _unquote_sheet_name(_quote_sheet_name(name)) == name

    def test_apostrophe_sheet_reference(self):
        """Produced '=John''s Data'!B5 — a syntactically broken formula."""
        result = adjust_formula_references(
            "='John''s Data'!T1.B[0]", 5, {}, {"John's Data": {"T1": 1}}
        )
        assert result == "='John''s Data'!B2"

    def test_ampersand_sheet_reference(self):
        result = adjust_formula_references(
            "='P&L'!T1.B[0]", 5, {}, {"P&L": {"T1": 1}}
        )
        assert result == "='P&L'!B2"

    def test_quoted_name_after_an_operator(self):
        """Quoting is what makes a hyphenated name unambiguous vs subtraction."""
        result = adjust_formula_references(
            "=A1-'FY-2024'!T1.B[0]", 5, {}, {"FY-2024": {"T1": 1}}
        )
        assert result == "=A1-'FY-2024'!B2"

    def test_quoted_cross_sheet_function(self):
        result = adjust_formula_references(
            "='Sales Data'!T1.SUM(B[0]:D[0])", 10, {}, {"Sales Data": {"T1": 3}}
        )
        assert result == "=SUM('Sales Data'!B4:D4)"

    def test_bare_names_unaffected(self):
        assert adjust_formula_references(
            "=Revenue!T1.B[0]", 5, {}, {"Revenue": {"T1": 1}}
        ) == "=Revenue!B2"

    def test_subtraction_from_bare_name_unaffected(self):
        assert adjust_formula_references(
            "=A1-Data!T1.B[0]", 5, {}, {"Data": {"T1": 1}}
        ) == "=A1-Data!B2"


class TestNonSpreadsheetNumbers:
    """float() accepts literals that have no place in a cell."""

    @pytest.mark.parametrize("raw", ["nan", "NaN", "inf", "-inf", "Infinity", "-Infinity"])
    def test_non_finite_stays_text(self, raw):
        """openpyxl writes these as an empty <v>, so the cell arrived blank."""
        result = resolve_cell(raw)
        assert result.value == raw
        assert isinstance(result.value, str)

    @pytest.mark.parametrize("raw", ["1_000", "1_000_000"])
    def test_underscore_literal_stays_text(self, raw):
        """PEP 515 underscores turned a product code into a number."""
        assert resolve_cell(raw).value == raw

    @pytest.mark.parametrize("raw,expected", [
        ("1234", 1234.0),
        ("-12.5", -12.5),
        ("0", 0.0),
    ])
    def test_ordinary_numbers_unaffected(self, raw, expected):
        assert resolve_cell(raw).value == pytest.approx(expected)

    def test_nan_cell_is_not_blank_in_output(self):
        wb = load_workbook(io.BytesIO(_workbook_bytes(
            "| Item | Value |\n|------|-------|\n| A | nan |\n"
        )))
        assert wb.active["B2"].value == "nan"


class TestInlineFormattingKeepsFontName:
    def test_bold_preserves_family(self):
        from openpyxl.styles import Font
        cell = Workbook().active["A1"]
        cell.font = Font(name="Arial", size=11)
        apply_cell_formatting(cell, {'bold': True, 'italic': False, 'monospace': False})
        assert cell.font.name == "Arial"
        assert cell.font.bold is True

    def test_italic_preserves_family(self):
        from openpyxl.styles import Font
        cell = Workbook().active["A1"]
        cell.font = Font(name="Arial", size=11)
        apply_cell_formatting(cell, {'bold': False, 'italic': True, 'monospace': False})
        assert cell.font.name == "Arial"
        assert cell.font.italic is True

    def test_monospace_overrides_family_deliberately(self):
        from openpyxl.styles import Font
        cell = Workbook().active["A1"]
        cell.font = Font(name="Arial", size=11)
        apply_cell_formatting(cell, {'bold': False, 'italic': False, 'monospace': True})
        assert cell.font.name == "Courier New"


class TestExcelTableHeaders:
    """Excel Tables need non-empty, unique column names or the file won't open."""

    def test_duplicate_headers_are_disambiguated(self):
        """Two 'Q1' columns produced a file Excel opened with a repair prompt."""
        from xlsx_tools.helpers import add_table_to_sheet
        wb = Workbook()
        add_table_to_sheet([["Q1", "Q1", "Q1"], ["1", "2", "3"]], wb.active, 1,
                           {}, {}, auto_filter=True)
        buf = io.BytesIO()
        wb.save(buf)
        names = _table_column_names(buf.getvalue())
        assert names == ["Q1", "Q1_2", "Q1_3"]
        assert len(set(names)) == len(names)

    def test_empty_header_gets_a_name(self):
        from xlsx_tools.helpers import add_table_to_sheet
        wb = Workbook()
        add_table_to_sheet([["Name", ""], ["a", "b"]], wb.active, 1,
                           {}, {}, auto_filter=True)
        buf = io.BytesIO()
        wb.save(buf)
        names = _table_column_names(buf.getvalue())
        assert names == ["Name", "Column2"]
        assert "" not in names

    def test_case_insensitive_duplicates_disambiguated(self):
        """Excel compares table column names case-insensitively."""
        from xlsx_tools.helpers import add_table_to_sheet
        wb = Workbook()
        add_table_to_sheet([["Total", "TOTAL"], ["1", "2"]], wb.active, 1,
                           {}, {}, auto_filter=True)
        buf = io.BytesIO()
        wb.save(buf)
        names = _table_column_names(buf.getvalue())
        assert len({n.casefold() for n in names}) == 2

    @pytest.mark.parametrize("headers,expected", [
        # The positional fallback collides with a real column of that name.
        (["Column2", ""], ["Column2", "Column2_2"]),
        # ...and with a later blank whose fallback matches an earlier header.
        (["Column3", "", ""], ["Column3", "Column2", "Column3_2"]),
        # A real header already shaped like a generated suffix.
        (["Q1", "Q1", "Q1_2"], ["Q1", "Q1_2", "Q1_2_2"]),
        (["", ""], ["Column1", "Column2"]),
    ])
    def test_fallback_names_are_deduplicated_too(self, headers, expected):
        """The blank-header fallback must clear the uniqueness check as well.

        Exempting it reintroduced the duplicate this function exists to
        prevent, whenever a real column was named "ColumnN" and column N
        happened to be blank.
        """
        from xlsx_tools.helpers import add_table_to_sheet
        wb = Workbook()
        add_table_to_sheet([headers, ["x"] * len(headers)], wb.active, 1,
                           {}, {}, auto_filter=True)
        buf = io.BytesIO()
        wb.save(buf)
        names = _table_column_names(buf.getvalue())
        assert names == expected
        assert len({n.casefold() for n in names}) == len(names)

    def test_unique_headers_untouched(self):
        from xlsx_tools.helpers import add_table_to_sheet
        wb = Workbook()
        add_table_to_sheet([["Region", "Amount"], ["N", "1"]], wb.active, 1,
                           {}, {}, auto_filter=True)
        buf = io.BytesIO()
        wb.save(buf)
        assert _table_column_names(buf.getvalue()) == ["Region", "Amount"]

    def test_headers_untouched_without_auto_filter(self):
        """No Table means no constraint — leave the user's headings alone."""
        wb = load_workbook(io.BytesIO(_workbook_bytes(
            "| Q1 | Q1 |\n|----|----|\n| 1 | 2 |\n"
        )))
        assert [wb.active["A1"].value, wb.active["B1"].value] == ["Q1", "Q1"]


class TestFormulaLengthGuard:
    def test_over_length_formula_stored_as_text(self):
        """An over-long formula made Excel reject the whole workbook."""
        long_formula = "=" + "+".join(f"A{i}" for i in range(1, 1600))
        assert len(long_formula) > MAX_FORMULA_LENGTH
        wb = load_workbook(io.BytesIO(_workbook_bytes(
            f"| Item | Value |\n|------|-------|\n| A | {long_formula} |\n"
        )))
        assert wb.active["B2"].value == long_formula
        assert wb.active["B2"].data_type == "s", "must not be written as a formula"

    def test_over_length_formula_warns(self, caplog):
        long_formula = "=" + "+".join(f"A{i}" for i in range(1, 1600))
        with caplog.at_level("WARNING"):
            _workbook_bytes(f"| Item | Value |\n|------|-------|\n| A | {long_formula} |\n")
        assert "over Excel's" in caplog.text

    def test_text_stored_formula_is_not_a_cycle(self):
        """A text-stored formula is never evaluated, so it can't be circular.

        The cell value still starts with '=', so a detector keyed on that
        alone would invent a cycle for a formula Excel treats as a label.
        """
        from xlsx_tools.circular_refs import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws["A1"] = "=A1+1"          # self-referential text, not a formula
        ws["A1"].data_type = "s"
        buf = io.BytesIO()
        wb.save(buf)
        assert detect_circular_references(buf.getvalue(), wb.sheetnames) == []

    def test_real_formula_still_detected_as_a_cycle(self):
        from xlsx_tools.circular_refs import detect_circular_references

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws["A1"] = "=A1+1"
        buf = io.BytesIO()
        wb.save(buf)
        assert len(detect_circular_references(buf.getvalue(), wb.sheetnames)) == 1

    def test_normal_formula_still_a_formula(self):
        wb = load_workbook(io.BytesIO(_workbook_bytes(
            "| Item | Value |\n|------|-------|\n| A | 1 |\n| T | =SUM(T1.B[0]:T1.B[0]) |\n"
        )))
        assert wb.active["B3"].value == "=SUM(B2:B2)"
        assert wb.active["B3"].data_type == "f"


class TestSharedBuildPath:
    """Both entry points must go through one implementation."""

    MARKDOWN = """## Sheet: Revenue
| Item | Value |
|------|-------|
| A    | 1500.75 |
| T    | =SUM(T1.B[0]:T1.B[0]) |
"""

    def test_buffer_and_upload_paths_agree(self):
        buffered = load_workbook(_markdown_to_excel_buffer(self.MARKDOWN))
        uploaded = load_workbook(io.BytesIO(_workbook_bytes(self.MARKDOWN)))
        assert buffered.sheetnames == uploaded.sheetnames
        for coord in ("A1", "B1", "B2", "B3"):
            assert buffered["Revenue"][coord].value == uploaded["Revenue"][coord].value
            assert (buffered["Revenue"][coord].number_format
                    == uploaded["Revenue"][coord].number_format)

    def test_buffer_path_detects_circular_references(self, caplog):
        """The LibreChat path must not miss the diagnostic."""
        with caplog.at_level("WARNING"):
            _markdown_to_excel_buffer(
                "| Item | Value |\n|------|-------|\n| A | =B2+1 |\n"
            )
        assert "circular" in caplog.text.lower()

    def test_buffer_path_warns_on_colliding_sheets(self, caplog):
        with caplog.at_level("WARNING"):
            _markdown_to_excel_buffer(
                "## Sheet: Report\n| A |\n|---|\n| 1 |\n\n"
                "## Sheet: Report\n| B |\n|---|\n| 2 |\n"
            )
        assert "collides" in caplog.text

    @pytest.mark.parametrize("bad_input", ["", "   ", "no tables here"])
    def test_both_paths_reject_the_same_input(self, bad_input):
        with pytest.raises(RuntimeError):
            _build_workbook(bad_input)
        with pytest.raises(RuntimeError):
            _markdown_to_excel_buffer(bad_input)

    def test_buffer_is_positioned_at_start(self):
        assert _markdown_to_excel_buffer(self.MARKDOWN).tell() == 0
